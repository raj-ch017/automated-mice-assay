import openpyxl
import pandas as pd
import os


def loading_data(sheet_name, file_location):
    """
    the function takes name of the sheet and file location as input
    and returns the active worksheet containing mouse trial data.
    the function prints a message that sheet has been successfully loaded
    and the number of rows and columns in the sheet.

    Input:

    - sheet_name (str): name of the sheet containing data
    - file_location (str): excel workboook file location

    Output:

    - work_sheet (openpyxl.worksheet): the active worksheet with mouse data  
    """

    work_book = openpyxl.load_workbook(file_location)
    work_sheet = work_book[sheet_name]

    row_num = work_sheet.max_row
    col_num = work_sheet.max_column
    print("Worksheet {} succesfully loaded.".format(sheet_name))
    print("Rows: {} | Columns: {}".format(row_num - 1, col_num))

    return work_sheet


def extracting_data(input_worksheet):
    """
    this function does the following things:

    - creates empty lists to store data from worksheet
    - the data is then extracted from the worksheet into appropriate lists
    - next, time interval is calculated for the following: interval from actvation from ROI1 to ROI 2 / ROI 3, interval from ROI 2 / ROI 3 activation to ROI 4 / ROI 5 activation, interval to reset trial
    - next, converting all timestamps from milliseconds to seconds
    - 
    """

    rows = input_worksheet.max_row
    cols = input_worksheet.max_column

    # list containing ROI counter data
    ROI_list = []
    # list containing trial start time data
    start_time = []
    # list containing checkpoint1 data 
    checkpoint1 = []
    # list containing decision-making time
    delta_t1 = []
    # list containing checkpoint2 
    checkpoint2 = []

    start_time = [input_worksheet.cell(row = x1, column = 1).value for x1 in range(2, rows + 1) if input_worksheet.cell(row = x1, column = 1).value is not None]
    ROI_list = [input_worksheet.cell(row = x2, column = 2).value for x2 in range(2, rows + 1) if input_worksheet.cell(row = x2, column = 2).value is not None ]
    checkpoint1 = [input_worksheet.cell(row = x3, column = 3).value for x3 in range(2, rows + 1) if input_worksheet.cell(row = x3, column = 3).value is not None]
    checkpoint2 = [input_worksheet.cell(row = x4, column = 4).value for x4 in range(2, rows + 1) if input_worksheet.cell(row = x4, column = 4).value is not None]

    data_dict = dict(StartTime = start_time, ROI = ROI_list, Checkpoint1 = checkpoint1, Checkpoint2 = checkpoint2)
    data_frame = pd.DataFrame(data_dict)


    return data_frame





def writing_to_excel(data_frame, file_details, name_of_sheet, count):

    if count == 0:
        with pd.ExcelWriter(file_details, engine='openpyxl') as writer:
            data_frame.to_excel(writer, sheet_name= name_of_sheet, index=False)
            count += 1

    else:
        with pd.ExcelWriter(file_details, engine = 'openpyxl', mode = 'a') as writer:

            data_frame.to_excel(writer, sheet_name= name_of_sheet, index=False)
    
    return count


def function_to_run(sheet_list, location_of_file, location_to_save):
    
    file_name = input("Enter name of the file: ") + ".xlsx"
    location_to_save = os.path.join(location_to_save, file_name)

    #list_of_df = []
    counter = 0

    for name_of_sheet in sheet_list:

        the_worksheet = loading_data(name_of_sheet, location_of_file)
        mice_df = extracting_data(the_worksheet)
        #list_of_df.append(mice_df)
        counter = writing_to_excel(mice_df, location_to_save, name_of_sheet, counter)


location_of_excel_save = r"C:\Users\Shrivas\Desktop\DM_RkP"
excel_workbook = r"C:\Users\Shrivas\Desktop\Copy of PLX-DAQ-v2.11.xlsm"
sheet_list = [  "WTM1_D14" ,"WTM2_D14" ,"WTM3_D14" , "WTC1F1_D14", "WTC1F2_D14", "WTC1F3_D14", "WTC1F4_D14", "WTC1F5_D14" , "WTC1F6_D14" , "VGluthm4diC1M1_D14" , "VGluthm4diC1M2_D14", "VGluthm4diC1M3_D14"  ]

function_to_run(sheet_list, excel_workbook, location_of_excel_save)