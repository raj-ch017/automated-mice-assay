import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os



def loading_data(sheet_name, file_location):
    """
    the function takes name of the sheet and file location as input
    and returns the active worksheet containing mouse trial data.
    the function prints a message that sheet has been successfully loaded
    and the number of rows and columns in the sheet.

    Input:

    - sheet_name (str): name of the sheet containing data
    - file_location (str): excel workboook file location

    Output:

    - work_sheet (openpyxl.worksheet): the active worksheet with mouse data  
    """

    work_book = openpyxl.load_workbook(file_location)
    work_sheet = work_book[sheet_name]

    row_num = work_sheet.max_row
    col_num = work_sheet.max_column
    print("Worksheet {} succesfully loaded.".format(sheet_name))
    print("Rows: {} | Columns: {}".format(row_num - 1, col_num))

    return work_sheet



def extracting_data(input_worksheet):
    """
    this function does the following things:

    - creates empty lists to store data from worksheet
    - the data is then extracted from the worksheet into appropriate lists
    - next, time interval is calculated for the following: interval from actvation from ROI1 to ROI 2 / ROI 3, interval from ROI 2 / ROI 3 activation to ROI 4 / ROI 5 activation, interval to reset trial
    - next, converting all timestamps from milliseconds to seconds
    - 
    """

    rows = input_worksheet.max_row
    cols = input_worksheet.max_column

    # list containing ROI counter data
    ROI_list = []
    # list containing trial start time data
    start_time = []
    # list containing checkpoint1 data 
    checkpoint1 = []
    # list containing decision-making time
    delta_t1 = []
    # list containing checkpoint2 
    checkpoint2 = []
    # list for total trial duration
    total_duration = []
    # list for restarting_duration
    reset_duration = []
    # list for duration of each individual trial
    each_trial = []

    start_time = [input_worksheet.cell(row = x1, column = 1).value for x1 in range(2, rows + 1) if input_worksheet.cell(row = x1, column = 1).value is not None]
    ROI_list = [input_worksheet.cell(row = x2, column = 2).value for x2 in range(2, rows + 1) if input_worksheet.cell(row = x2, column = 2).value is not None ]
    checkpoint1 = [input_worksheet.cell(row = x3, column = 3).value for x3 in range(2, rows + 1) if input_worksheet.cell(row = x3, column = 3).value is not None]
    checkpoint2 = [input_worksheet.cell(row = x4, column = 4).value for x4 in range(2, rows + 1) if input_worksheet.cell(row = x4, column = 4).value is not None]


    delta_t1, total_duration, reset_duration, each_trial = calculating_duration(start_time, checkpoint1, checkpoint2)

    start_time, checkpoint1, delta_t1, checkpoint2, total_duration, reset_duration, each_trial = convert_to_seconds(start_time, checkpoint1, delta_t1, checkpoint2, total_duration, reset_duration, each_trial)

    return (start_time, ROI_list, checkpoint1, delta_t1, checkpoint2, total_duration, reset_duration, each_trial)
    




def choice_percentage(roi_counter):
    """
    calculates and prints the number of times ROI 2 and ROI 3 are activated.

    Input:

    - roi_counter (list): A list containing ROI entries, where each element
        is an integer representing activation of either ROI 2 or ROI 3

    Output:

    - tuple containing ROI 2 count and ROI 3 count
    """

    roi2_count = 0
    roi3_count = 0

    for ele in roi_counter:
        
        if ele == 2:
            roi2_count += 1
        elif ele == 3:
            roi3_count += 1
    
    print("ROI2: {} | ROI3: {}".format(roi2_count, roi3_count))
    return (roi2_count, roi3_count)



def convert_to_seconds(list1, list2, list3, list4, list5, list6, list7):
    """
    converts multiple lists of time values from milliseconds to seconds.

    each value in the input lists is millisecond time data and the function 
    converts the time to seconds by dividing by 1000.

    Inputs:

    - list1: list containing start time in milliseconds
    - list2: list containing time at which ROI2 / ROI3 was activated
    - list3: list containing duration to cross ROI2 / ROI3 after crossing ROI1 
    - list4: list containing time at which ROI4 / ROI5 was activated
    - list5: list containing duration between ROI2 / ROI3 activation and ROI4 / ROI5 activation
    - list6: list containing duration between ROI4 / ROI5 activation and ROI1 (time aken to restart trial)

    Output:

    tuple of lists - a tuple containing six lists, each corresponding to the converted input lists, with all values in seconds
    """

    output_list1 = []
    output_list2 = []
    output_list3 = []
    output_list4 = []
    output_list5 = []
    output_list6 = []
    output_list7 = []

    condition = (len(list1) == len(list2) == len(list3) == len(list4) == len(list5))
    
    assert (condition == True)

    for k in range(len(list1)):
        
        #print("i am doing this shit")
        output_list1.append(list1[k] / 1000)
        output_list2.append(list2[k] / 1000)
        output_list3.append(list3[k] / 1000)
        output_list4.append(list4[k] / 1000)
        output_list5.append(list5[k] / 1000)
        output_list6.append(list6[k] / 1000)
        output_list7.append(list7[k] / 1000)

    return (output_list1, output_list2, output_list3, output_list4, output_list5, output_list6, output_list7)


def calculating_duration(start_time, data_series1, data_series2):
    """
    calculates duration between sequential time checkpoints for multiple trials

    the function calculates the following duration for each trial:
        - time from activation of ROI1 (start_time) to activation of ROI2 / ROI3 (data_series1)
        - time from activation of ROI2 / ROI3 (data_series1) to activation of ROI4 / ROI5 (data_series2)
        - time from activation of ROI4 / ROI5 (data_series2) to next activation of ROI1 

    Input:

    - start_time: list of time containing starting timestamps for each trial
    - data_series: list of timestamps for activation of ROI2 / ROI3
    - data_series: list of timestamps for activation of ROI4 / ROI5

    Output:

    - tuple of list containing the following data: 
        - checkpoint1_duration: list of duration from start_time to data_series1
        - checkpoint2_duration: list of duration from data_series1 to data_series2
        - checkpoint3_duration: list of duration from data_series2 to start_time of next trial
    """

    checkpoint1_duration = []
    checkpoint2_duration = []
    checkpoint3_duration = []
    individual_trial = []

    assert len(start_time) == len(data_series1) == len(data_series2),"Length Mismatch"

    for ind in range(len(data_series1)):

        t1 = start_time[ind]
        t2 = data_series1[ind]
        t3 = data_series2[ind]

        checkpoint1_duration.append(t2 - t1)
        checkpoint2_duration.append(t3 - t2)
        individual_trial.append(t3-t1)

        if ind != len(start_time) - 1:
            t4 = start_time[ind+1]
            checkpoint3_duration.append(t4 - t3)
        else:
            checkpoint3_duration.append(0)



    return (checkpoint1_duration, checkpoint2_duration, checkpoint3_duration, individual_trial)

    



def saving_data(start_time, ROI_list, checkpoint1, delta_t1, checkpoint2, total_duration, reset_time, mice_name, each_time):

    data_dict = dict(ROI = ROI_list, StartTime = start_time, Checkpoint1 = checkpoint1, DecisionTime = delta_t1, Checkpoint2 = checkpoint2, ArmDuration = total_duration, RestartTime = reset_time, TrialDuration = each_time)
    data_frame = pd.DataFrame(data_dict)


    #assert mice_name.isalnum()

    data_frame.to_csv(mice_name + '_mice_data.csv')


    x_values = [k+1 for k in range(len(ROI_list))]

    # Creating plot for DecisionTime
    plt.bar(x_values, data_frame['DecisionTime'], color = 'red')
    plt.locator_params(integer=True)
    plt.title("Decision Latency")
    plt.xlabel('Trial Number')
    plt.ylabel('Time (s)')
    plt.savefig(mice_name + '_decision_time.jpg')
    plt.close()

    # Creating plot for time spent in arm
    plt.bar(x_values, data_frame['ArmDuration'], color = 'green')
    plt.locator_params(integer=True)
    plt.title("Time Spent in Arm")
    plt.xlabel('Trial Number')
    plt.ylabel('Time (s)')
    plt.savefig(mice_name + '_time_in_arm.jpg')
    plt.close()

    # Creating plot for resetting trial
    plt.bar(x_values[:-1], data_frame['RestartTime'][:-1], color = 'blue')
    plt.locator_params(integer=True)
    plt.title("Trial Restart Time")
    plt.xlabel('Trial Number')
    plt.ylabel('Time (s)')
    plt.savefig(mice_name + '_trial_restart.jpg')
    plt.close()

    # Creating plot for trial duration
    plt.bar(x_values, data_frame['TrialDuration'], color = 'magenta')
    plt.locator_params(integer = True)
    plt.title("Trial Duration")
    plt.xlabel('Trial Number')
    plt.ylabel('Time (s)')
    plt.savefig(mice_name + '_trial_duration.jpg')
    plt.close()

    # getting ROI choice percentage
    ROI2, ROI3 = choice_percentage(ROI_list)

    # Creating pie-chart to represent ROI choice
    plt.pie([ROI2, ROI3], labels = [('ROI2', ROI2), ('ROI3', ROI3)], autopct='%1.1f%%')
    plt.title("ROI Choice")
    plt.savefig(mice_name + '_ROISelection.jpg')
    plt.close()

    # Creating boxplot for time intervals
    boxplot_data = [data_frame['DecisionTime'], data_frame['ArmDuration'], data_frame['RestartTime'][:-1], data_frame['TrialDuration']]
    fig = plt.figure(figsize =(7, 7))
    ax = fig.add_axes([0, 0, 1, 1])
    label_list = ['DecisionTime', 'ArmTime', 
                    'ResetTime']
    ax.set_yticklabels(['DecisionTime', 'ArmTime', 
                    'ResetTime', 'TrialDuration'])
    bp = ax.boxplot(boxplot_data, meanline=True, vert=False, showmeans = True, medianprops={'linewidth': 2, 'color': 'red'}, meanprops={'linewidth': 2, 'color': 'orange'}, patch_artist=True)
    plt.legend([bp['medians'][0], bp['means'][0]], ['median', 'mean'])
    plt.title("BoxPlot for Time Intervals")
    plt.savefig(mice_name + '_boxplot.jpg', bbox_inches = 'tight')
    plt.close() 

    
    

def creating_folder(mice_name, old_location, start_time,  roi_active, activation_roi, t1_delta, last_roi_active, time_arm, time_reset, time_for_each):

    os.chdir(old_location)
    #mice_name = input("Please enter mice name: ")
    #assert mice_name.isalnum()
    new_location = old_location + "\\" + mice_name

    os.mkdir(mice_name)
    print("{} folder successfully created.".format(mice_name))
    os.chdir(new_location)
    saving_data(start_time, roi_active, activation_roi, t1_delta, last_roi_active, time_arm, time_reset, mice_name, time_for_each)
    os.chdir(old_location)

def create_data_folder(old_location):

    folder_name = input("Enter the name of the folder: ")
    #assert folder_name.isalnum()
    new_location = os.path.join(old_location, folder_name)
    os.mkdir(new_location)
    return new_location
    



def master_function(name_of_sheet, file_location, data_store):
    """
    the function takes in location of the excel file and location
    of folder where processed data and graph for each mice will be stored

    Input:

    - file_location (str): string datatype containing location of excel sheet
    - data_store (str): string datatype containing location of mice folder

    Output:

        TRUE
    """


    #name_of_sheet = input("Please enter name of the sheet: ")

    #assert(name_of_sheet.isalnum())

    working_sheet = loading_data(name_of_sheet, file_location)

    start_time, roi_active, activation_roi, t1_delta, last_roi_active, time_arm, time_reset, each_trial = extracting_data(working_sheet)

    creating_folder(name_of_sheet, data_store, start_time, roi_active, activation_roi, t1_delta, last_roi_active, time_arm, time_reset, each_trial)

    return True

def run_this_function(data_list, file_location, data_store_loc):

    for ind in range(len(data_list)):

        sheet_name = data_list[ind]

        master_function(sheet_name, file_location, data_store_loc)

### RUNTIME START ###

# location of excel sheet containing data
location_of_excel = r"C:\Users\Shrivas\Desktop\DM_RkP\day6_freechoice_CBC.xlsx"
#sheet_list = ["Tac1hm3dqC1M1", "Tac1hm3dqC1M2", "Tac1hm3dqC1M3", "Tac1hm3dqC2M1", "Tac1hm3dqC2M2", "Tac1hm3dqC2M3", "VGluthm4diC1M1", "VGluthm4diC1M2", "VGluthm4diC1M3", "WTM1", "WTM2", "WTM3","Tac1hm3dqC1F1", "Tac1hm3dqC1F2", "Tac1hm3dqC1F3", "Tac1hm3dqC2F1", "Tac1hm3dqC2F2", "Tac1hm3dqC2F3", "WTC1F1", "WTC1F2", "WTC1F3", "WTC1F4", "WTC1F5", "WTC1F6"   ]
#sheet_list = ["Tachm3dqC1M1", "Tachm3dqC1M2", "Tachm3dqC1M3", "Tachm3dqC2M1", "Tachm3dqC2M2", "Tachm3dqC2M3", "VGluthm4diC1M1", "VGluthm4diC1M2", "VGluthm4diC1M3", "WTM1", "WTM2", "WTM3","Tachm3dqC1F1", "Tachm3dqC1F2", "Tachm3dqC1F3", "Tachm3dqC2F1", "Tachm3dqC2F2", "Tachm3dqC2F3", "WTC1F1", "WTC1F2", "WTC1F3", "WTC1F4", "WTC1F5", "WTC1F6"   ]
sheet_list = [ "Tac1hm3dqC1M1_D6", "Tac1hm3dqC1M2_D6", "Tac1hm3dqC1M3_D6", "Tac1hm3dqC2M1_D6", "Tac1hm3dqC2M2_D6", "Tac1hm3dqC2M3_D6","Tac1hm3dqC1F1_D6", "Tac1hm3dqC1F2_D6", "Tac1hm3dqC1F3_D6", "Tac1hm3dqC2F1_D6", "Tac1hm3dqC2F2_D6", "Tac1hm3dqC2F3_D6"]

# location where processed data and graphs will be stored
store_loc = r"C:\Users\Shrivas\Desktop\DM_RkP"


# master function to run the show
storage_location = create_data_folder(store_loc)
run_this_function(sheet_list, location_of_excel, storage_location)

